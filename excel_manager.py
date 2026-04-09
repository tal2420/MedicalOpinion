import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config import get_data_dir
import schema

EXCEL_FILENAME = "מאגר_תיקים.xlsx"
SHEET_NAME = "תיקים"

# Color palette for status cells (hex without #)
COLOR_MAP = {
    "success": "C6EFCE",
    "warning": "FFEB9C",
    "danger": "FFC7CE",
    "info": "B4C6E7",
}


def get_excel_path():
    return os.path.join(get_data_dir(), EXCEL_FILENAME)


def _create_workbook():
    """Create a new formatted Excel workbook using schema definitions."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_view.rightToLeft = True

    columns = schema.get_columns()
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.freeze_panes = "A2"

    # Store schema version in a custom property-like cell
    meta_ws = wb.create_sheet("_meta")
    meta_ws.cell(row=1, column=1, value="schema_version")
    meta_ws.cell(row=1, column=2, value=schema.SCHEMA_VERSION)
    meta_ws.sheet_state = "hidden"

    wb.save(get_excel_path())
    return wb


def _migrate_workbook(wb):
    """Add any missing columns to an existing workbook (non-destructive).

    Compares the existing header row against the current schema and appends
    any new columns that don't exist yet. Never removes or reorders columns.
    """
    ws = wb[SHEET_NAME]
    columns = schema.get_columns()

    # Read existing headers
    existing_headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            existing_headers.append(val)

    existing_labels = set(existing_headers)

    # Find new columns that need to be added
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    added = 0
    for label, width in columns:
        if label not in existing_labels:
            new_col = ws.max_column + 1
            cell = ws.cell(row=1, column=new_col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(new_col)].width = width
            added += 1

    if added > 0:
        # Update filter range
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"

        # Update schema version
        if "_meta" not in wb.sheetnames:
            meta_ws = wb.create_sheet("_meta")
            meta_ws.sheet_state = "hidden"
        else:
            meta_ws = wb["_meta"]
        meta_ws.cell(row=1, column=1, value="schema_version")
        meta_ws.cell(row=1, column=2, value=schema.SCHEMA_VERSION)

        wb.save(get_excel_path())
        print(f"[excel_manager] Migrated: added {added} new column(s)")

    return wb


def _get_header_map(ws):
    """Build a mapping from field key -> column index based on actual headers.

    This makes reading resilient to column reordering or extra columns.
    """
    labels_to_key = {f["label"]: f["key"] for f in schema.get_all_fields()}
    header_map = {}  # key -> col_idx (1-based)

    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col).value
        if header_val and header_val in labels_to_key:
            header_map[labels_to_key[header_val]] = col

    return header_map


def _get_workbook():
    """Get or create the Excel workbook. Runs migration if needed."""
    path = get_excel_path()
    if os.path.exists(path):
        wb = load_workbook(path)
        if SHEET_NAME not in wb.sheetnames:
            return _create_workbook()
        return _migrate_workbook(wb)
    return _create_workbook()


def get_next_case_number():
    """Get the next auto-incremented case number."""
    wb = _get_workbook()
    ws = wb[SHEET_NAME]
    header_map = _get_header_map(ws)
    cn_col = header_map.get("case_number", 1)
    max_num = 0
    for row in ws.iter_rows(min_row=2, min_col=cn_col, max_col=cn_col, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            max_num = max(max_num, int(row[0]))
    return max_num + 1


def add_case(data):
    """Add a new case row. Returns the case number."""
    wb = _get_workbook()
    ws = wb[SHEET_NAME]
    header_map = _get_header_map(ws)

    case_number = get_next_case_number()
    data["case_number"] = case_number

    # Apply defaults from schema for missing fields
    defaults = schema.get_defaults()
    for key, default_val in defaults.items():
        if key not in data or not data[key]:
            data[key] = default_val

    if not data.get("date_received"):
        data["date_received"] = datetime.now().strftime("%d/%m/%Y")

    new_row = ws.max_row + 1
    cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for key, col_idx in header_map.items():
        value = data.get(key, "")
        cell = ws.cell(row=new_row, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border
        cell.font = Font(name="Arial", size=10)

    _apply_status_colors(ws, new_row, header_map)
    wb.save(get_excel_path())
    return case_number


def _apply_status_colors(ws, row, header_map):
    """Apply color coding to status cells based on schema definitions."""
    status_colors = schema.get_status_colors()
    status_fields = schema.get_status_fields()

    for key in status_fields:
        if key not in header_map:
            continue
        col_idx = header_map[key]
        cell = ws.cell(row=row, column=col_idx)
        val = str(cell.value) if cell.value else ""
        color_name = status_colors.get(val)
        if color_name and color_name in COLOR_MAP:
            hex_color = COLOR_MAP[color_name]
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def update_case(case_number, data):
    """Update fields for an existing case."""
    wb = _get_workbook()
    ws = wb[SHEET_NAME]
    header_map = _get_header_map(ws)
    cn_col = header_map.get("case_number", 1)

    target_row = None
    for row in ws.iter_rows(min_row=2, min_col=cn_col, max_col=cn_col):
        if row[0].value and int(row[0].value) == int(case_number):
            target_row = row[0].row
            break

    if not target_row:
        raise ValueError(f"Case {case_number} not found")

    for key, value in data.items():
        if key in header_map and key != "case_number":
            ws.cell(row=target_row, column=header_map[key], value=value)

    _apply_status_colors(ws, target_row, header_map)
    wb.save(get_excel_path())


def _backfill_first_last_name(case):
    """If case has plaintiff_name but no first/last, split on the fly.

    Also normalizes plaintiff_name to "first last" (with space) when it was
    stored as concatenated text like "אילןגרינולד".
    """
    full = (case.get("plaintiff_name") or "").strip()
    has_first = bool(case.get("plaintiff_first_name"))
    has_last = bool(case.get("plaintiff_last_name"))

    if not full and not has_first and not has_last:
        return

    try:
        from email_parser import _split_name

        # If we have first/last but no full name, build it
        if (has_first or has_last) and not full:
            case["plaintiff_name"] = " ".join(
                p for p in [case.get("plaintiff_first_name", ""), case.get("plaintiff_last_name", "")] if p
            )
            return

        # If we have a full name, derive first/last from it
        first, last = _split_name(full)
        if not has_first and first:
            case["plaintiff_first_name"] = first
        if not has_last and last:
            case["plaintiff_last_name"] = last

        # Normalize the displayed full name with a space
        if first and last and " " not in full:
            case["plaintiff_name"] = f"{first} {last}"
    except Exception:
        pass


def get_all_cases():
    """Return all cases as list of dicts."""
    wb = _get_workbook()
    ws = wb[SHEET_NAME]
    header_map = _get_header_map(ws)
    # Invert: col_idx -> key
    col_to_key = {v: k for k, v in header_map.items()}

    cases = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        # Skip empty rows
        cn_col = header_map.get("case_number", 1)
        if not row[cn_col - 1].value:
            continue
        case = {}
        for col_idx, key in col_to_key.items():
            val = row[col_idx - 1].value
            case[key] = val if val is not None else ""
        # Ensure all schema keys exist (for new fields added after data was created)
        for field_key in schema.get_all_keys():
            if field_key not in case:
                case[field_key] = ""
        _backfill_first_last_name(case)
        cases.append(case)
    return cases


def get_case(case_number):
    """Return a single case by number."""
    wb = _get_workbook()
    ws = wb[SHEET_NAME]
    header_map = _get_header_map(ws)
    col_to_key = {v: k for k, v in header_map.items()}
    cn_col = header_map.get("case_number", 1)

    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[cn_col - 1].value and int(row[cn_col - 1].value) == int(case_number):
            case = {}
            for col_idx, key in col_to_key.items():
                val = row[col_idx - 1].value
                case[key] = val if val is not None else ""
            for field_key in schema.get_all_keys():
                if field_key not in case:
                    case[field_key] = ""
            _backfill_first_last_name(case)
            return case
    return None


def delete_case(case_number):
    """Delete a case row."""
    wb = _get_workbook()
    ws = wb[SHEET_NAME]
    header_map = _get_header_map(ws)
    cn_col = header_map.get("case_number", 1)

    for row in ws.iter_rows(min_row=2, min_col=cn_col, max_col=cn_col):
        if row[0].value and int(row[0].value) == int(case_number):
            ws.delete_rows(row[0].row)
            wb.save(get_excel_path())
            return True
    return False


def get_dashboard_stats():
    """Return summary statistics for the dashboard, driven by schema."""
    cases = get_all_cases()
    counters = schema.get_dashboard_counters()

    stats = {}
    for counter in counters:
        name = counter["name"]
        if counter["field_key"] is None:
            # Total counter
            stats[name] = len(cases)
        else:
            field_key = counter["field_key"]
            count_values = counter["count_values"]
            stats[name] = sum(1 for c in cases if c.get(field_key) in count_values)

    return stats
